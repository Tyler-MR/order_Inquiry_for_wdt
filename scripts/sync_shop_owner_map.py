"""Export the Windows workbook and push a complete snapshot to the Linux host."""

from __future__ import annotations

import argparse
import json
import os
import shlex
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def read_workbook(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("缺少 openpyxl，请先执行：python -m pip install openpyxl") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {str(value).strip(): index for index, value in enumerate(header_row) if value is not None}
        shop_column = headers.get("店铺名称")
        owner_column = headers.get("负责人")
        if shop_column is None or owner_column is None:
            raise ValueError("Excel 必须包含“店铺名称”和“负责人”两列")

        platform_column = headers.get("平台")
        rows: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            shop_name = str(row[shop_column] or "").strip()
            owner_name = str(row[owner_column] or "").strip()
            platform = str(row[platform_column] or "").strip() if platform_column is not None else ""
            if shop_name or owner_name:
                if not shop_name or not owner_name:
                    raise ValueError("Excel 中存在只有店铺名称或只有负责人的不完整行")
                rows.append({"shop_name": shop_name, "owner_name": owner_name, "platform": platform})
        if not rows:
            raise ValueError("Excel 中没有有效的店铺负责人数据")
        return rows
    finally:
        workbook.close()


def write_snapshot(rows: list[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    payload: dict[str, Any] = {
        "version": datetime.now(timezone.utc).isoformat(),
        "items": rows,
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def run(command: list[str], *, label: str) -> None:
    print(f"[{label}] {' '.join(command)}")
    subprocess.run(command, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync shop-owner mapping to Linux")
    parser.add_argument("--excel", type=Path, default=Path("private/shop_owner_map.xlsx"))
    parser.add_argument("--ssh-key", type=Path, default=Path.home() / ".ssh" / "codex_linux_deploy_ed25519")
    parser.add_argument("--host", default=os.getenv("WDT_SYNC_SSH_HOST", "192.168.16.54"))
    parser.add_argument("--user", default=os.getenv("WDT_SYNC_SSH_USER", "lingchi"))
    parser.add_argument("--project", default=os.getenv("WDT_SYNC_REMOTE_PROJECT", "/home/lingchi/order_Inquiry_for_wdt"))
    parser.add_argument("--service", default="backend")
    parser.add_argument("--output", type=Path, default=Path("sync/shop_owner_map.json"))
    parser.add_argument("--json-only", action="store_true", help="只生成 JSON，不连接 Linux")
    args = parser.parse_args()

    excel = args.excel.resolve()
    if not excel.is_file():
        raise SystemExit(f"找不到 Excel：{excel}")
    if not args.ssh_key.is_file():
        raise SystemExit(f"找不到 SSH 私钥：{args.ssh_key}")

    rows = read_workbook(excel)
    write_snapshot(rows, args.output)
    if args.json_only:
        print(f"已生成 {args.output}：{len(rows)} 条")
        return

    remote = f"{args.user}@{args.host}"
    remote_sync = f"{args.project}/sync"
    remote_json = f"{remote_sync}/shop_owner_map.json"
    ssh_base = ["ssh", "-i", str(args.ssh_key), "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", remote]
    scp_base = ["scp", "-i", str(args.ssh_key), "-o", "BatchMode=yes", "-o", "ConnectTimeout=15"]

    run(ssh_base + [f"mkdir -p {shlex.quote(remote_sync)}"], label="prepare")
    run(scp_base + [str(args.output), f"{remote}:{remote_json}"], label="upload")
    import_command = (
        f"cd {shlex.quote(args.project)} && "
        f"docker compose exec -T {shlex.quote(args.service)} "
        f"python scripts/import_shop_owner_map.py /app/sync/shop_owner_map.json"
    )
    run(ssh_base + [import_command], label="import")
    print(f"同步完成：{len(rows)} 条")


if __name__ == "__main__":
    try:
        main()
    except subprocess.CalledProcessError as error:
        raise SystemExit(error.returncode) from error
