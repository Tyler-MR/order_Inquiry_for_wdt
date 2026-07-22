"""Read product workbooks on Windows and push a complete snapshot to Linux."""

from __future__ import annotations

import argparse
import json
import math
import os
import shlex
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


DEFAULT_SPEC_EXCEL = Path(r"C:\Users\Financial\Desktop\规格数量.xlsx")
DEFAULT_PRODUCT_EXCEL = Path(r"C:\Company\Python导入数据\表\产品成本佣金表.xlsx")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _headers(row: tuple[Any, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(row):
        header = _text(value).replace("\ufeff", "")
        if header:
            result.setdefault(header, index)
    return result


def read_product_names(path: Path) -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook

    records: dict[str, dict[str, str]] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header = _headers(next(rows))
            except StopIteration:
                continue
            code_column = header.get("商品编码")
            name_column = header.get("商品名称")
            if code_column is None or name_column is None:
                continue
            for row in rows:
                code = _text(row[code_column] if code_column < len(row) else None)
                name = _text(row[name_column] if name_column < len(row) else None)
                if not code or not name:
                    continue
                # 同一编码重复时保留 Excel 最后一次出现的名称。
                records[code] = {
                    "product_code": code,
                    "product_name": name,
                    "source_sheet": sheet.title,
                }
    finally:
        workbook.close()
    return records


def read_product_specs(path: Path) -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook

    records: dict[str, dict[str, str]] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = _headers(next(rows))
        except StopIteration:
            return records
        sku_column = header.get("sku编码")
        spec_column = header.get("产品规格")
        if sku_column is None or spec_column is None:
            raise ValueError("规格数量.xlsx 必须包含“sku编码”和“产品规格”两列")
        for row in rows:
            sku = _text(row[sku_column] if sku_column < len(row) else None)
            spec = _text(row[spec_column] if spec_column < len(row) else None)
            if not sku or not spec:
                continue
            # 同一 SKU 重复时保留 Excel 最后一次出现的规格。
            records[sku] = {"sku_code": sku, "product_spec": spec, "source_sheet": sheet.title}
    finally:
        workbook.close()
    return records


def build_items(spec_excel: Path, product_excel: Path) -> list[dict[str, str]]:
    product_names = read_product_names(product_excel)
    product_specs = read_product_specs(spec_excel)
    merged: dict[str, dict[str, str]] = {}

    for code, item in product_names.items():
        merged[code] = {
            "lookup_code": code,
            "product_code": item["product_code"],
            "sku_code": "",
            "product_name": item["product_name"],
            "product_spec": "",
            "source_sheet": item["source_sheet"],
            "source_file": product_excel.name,
        }

    for code, item in product_specs.items():
        current = merged.setdefault(
            code,
            {
                "lookup_code": code,
                "product_code": "",
                "sku_code": "",
                "product_name": "",
                "product_spec": "",
                "source_sheet": "",
                "source_file": "",
            },
        )
        current["sku_code"] = item["sku_code"]
        current["product_spec"] = item["product_spec"]
        current["source_sheet"] = ";".join(
            value for value in (current["source_sheet"], item["source_sheet"]) if value
        )
        current["source_file"] = ";".join(
            value for value in (current["source_file"], spec_excel.name) if value
        )

    return [merged[key] for key in sorted(merged)]


def write_snapshot(items: list[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": datetime.now(timezone.utc).isoformat(),
        "items": items,
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def run(command: list[str], *, label: str) -> None:
    print(f"[{label}] {' '.join(command)}")
    subprocess.run(command, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync product master data to Linux MySQL")
    parser.add_argument("--spec-excel", type=Path, default=DEFAULT_SPEC_EXCEL)
    parser.add_argument("--product-excel", type=Path, default=DEFAULT_PRODUCT_EXCEL)
    parser.add_argument("--ssh-key", type=Path, default=Path.home() / ".ssh" / "codex_linux_deploy_ed25519")
    parser.add_argument("--host", default=os.getenv("WDT_SYNC_SSH_HOST", "192.168.16.54"))
    parser.add_argument("--user", default=os.getenv("WDT_SYNC_SSH_USER", "lingchi"))
    parser.add_argument("--project", default=os.getenv("WDT_SYNC_REMOTE_PROJECT", "/home/lingchi/order_Inquiry_for_wdt"))
    parser.add_argument("--service", default="backend")
    parser.add_argument("--output", type=Path, default=Path("sync/product_master.json"))
    parser.add_argument("--json-only", action="store_true")
    args = parser.parse_args()

    spec_excel = args.spec_excel.resolve()
    product_excel = args.product_excel.resolve()
    if not spec_excel.is_file():
        raise SystemExit(f"找不到规格 Excel：{spec_excel}")
    if not product_excel.is_file():
        raise SystemExit(f"找不到商品名称 Excel：{product_excel}")
    if not args.json_only and not args.ssh_key.is_file():
        raise SystemExit(f"找不到 SSH 私钥：{args.ssh_key}")

    items = build_items(spec_excel, product_excel)
    if not items:
        raise SystemExit("两个 Excel 中没有有效的商品编码、SKU、名称或规格数据")
    write_snapshot(items, args.output)
    print(f"已生成商品基础数据：{args.output}，共 {len(items)} 条")
    if args.json_only:
        return

    remote = f"{args.user}@{args.host}"
    remote_sync = f"{args.project}/sync"
    remote_json = f"{remote_sync}/product_master.json"
    ssh_base = ["ssh", "-i", str(args.ssh_key), "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", remote]
    scp_base = ["scp", "-i", str(args.ssh_key), "-o", "BatchMode=yes", "-o", "ConnectTimeout=15"]
    run(ssh_base + [f"mkdir -p {shlex.quote(remote_sync)}"], label="prepare")
    run(scp_base + [str(args.output), f"{remote}:{remote_json}"], label="upload")
    import_command = (
        f"cd {shlex.quote(args.project)} && "
        f"docker compose exec -T {shlex.quote(args.service)} "
        f"python scripts/import_product_master.py /app/sync/product_master.json"
    )
    run(ssh_base + [import_command], label="import")
    print(f"同步完成：{len(items)} 条商品基础数据")


if __name__ == "__main__":
    main()
